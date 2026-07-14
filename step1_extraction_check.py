"""
VeloNEMO v2 — Step 1: can an LLM reproduce the human v1 metric extraction?

For each ground-truth paper: fetch OA PDF -> parse (drop references) -> Gemini extracts the
index's metrics against the ontology vocabulary (velonemo.py) with K-sample self-consistency
-> compare leniently to the human GT (context/metrics.xlsx). Output: step1/report/step1_report.md.

The LLM gets the existing concept vocabulary and reuses names where they fit, coining new ones
(is_new) only for genuinely new metrics (D-021). Comparison is per paper; no cross-paper ER (D-016).
Idempotent (cached; FORCE_EXTRACT to redo). Blocked PDFs: drop at step1/pdfs/<openalex_id>.pdf.
Run: python step1_extraction_check.py
"""
import csv
import json
import re
import time
import tomllib
import urllib.request
from collections import Counter, defaultdict
from difflib import SequenceMatcher
from pathlib import Path
from typing import List, Optional

import openpyxl
from pydantic import BaseModel
from google import genai
from google.genai import types

from velonemo import THEMATIC_TYPES, CRITERIA, FEATURES, SCALES, METRICS_BY_TYPE

# existing metric concepts grouped by thematic type — given to the LLM so it reuses established
# names (as the human annotator did) and only mints a new name for a genuinely new metric.
METRIC_VOCAB = "\n".join(f"  {t}: {', '.join(ms)}" for t, ms in METRICS_BY_TYPE.items())

ROOT = Path(__file__).resolve().parent
GROUND_TRUTH_XLSX = ROOT / "context" / "metrics.xlsx"     # human v1 ground truth (in context/)
MANIFEST_JSONL = ROOT / "corpus" / "manifest.jsonl"   # superset: DOI+pdf_url for all GT papers
STEP1 = ROOT / "step1"
for sub in ("pdfs", "parsed", "extracted", "report"):
    (STEP1 / sub).mkdir(parents=True, exist_ok=True)

cfg = tomllib.load(open(ROOT / "config.toml", "rb"))
client = genai.Client(api_key=cfg["keys"]["gemini_api_key"])
MODEL = "gemini-3.1-flash-lite"          # same model as corpus screening
EMBED_MODEL = "gemini-embedding-001"     # semantic definition matching (ER preview, D-005)
K, TEMP = 3, 0.4                         # self-consistency: K samples/paper, majority-vote (D-012)
MIN_INTERVAL = 60.0 / 60
FORCE_EXTRACT = False
NAME_FUZZ, DEF_FUZZ, SEM_FUZZ = 0.82, 0.60, 0.78   # name / def-lexical / def-semantic thresholds

SLICE_IDS = {                             # openalex_id -> short key (15 OA ground-truth papers)
    "W4353080938": "beecham2023",   "W3207880534": "hardinghaus2021",
    "W4319011370": "hsu2023",       "W2220486709": "krenn2015",
    "W3118359258": "schmidquerg2021", "W4361011429": "weikl2023",
    "W4291123807": "santos2022",    "W3161486144": "ito2021",
    "W2901426564": "abad2018",      "W2911593481": "daraei2021",
    "W2941312659": "boisjoly2019",  "W3169192492": "reggiani2021",
    "W4220966335": "soltani2022",   "W4226421486": "karolemeas2022",
    "W4310059561": "codina2022",
}
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                    "(KHTML, like Gecko) Chrome/120.0 Safari/537.36",
      "Accept": "application/pdf,text/html,*/*;q=0.8", "Accept-Language": "en-US,en;q=0.9"}


# --------------------------------------------------------------------------- helpers
def norm_doi(d):
    return re.sub(r"^https?://(dx\.)?doi\.org/", "", d.strip().lower()) if d else None


def key(s):                               # comparable metric key: alphanumerics only
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


# metric name -> the thematic type it is a subclass of IN THE ONTOLOGY (the canonical classification).
# lets us check the LLM's thematic_type against the ontology (not only the human GT row), and flag
# metrics where the GT row's type disagrees with the ontology.
ONTO_TYPE = {key(m): t for t, ms in METRICS_BY_TYPE.items() for m in ms}


def sim(a, b):
    return SequenceMatcher(None, a or "", b or "").ratio()


_emb = {}                                 # text -> embedding cache (semantic definition matching)
def embed(texts):
    todo = [t for t in dict.fromkeys(texts) if t and t not in _emb]
    for i in range(0, len(todo), 100):
        batch = todo[i:i + 100]
        for t, e in zip(batch, client.models.embed_content(model=EMBED_MODEL, contents=batch).embeddings):
            _emb[t] = e.values


def cos(a, b):
    if not a or not b:
        return 0.0
    return sum(x * y for x, y in zip(a, b)) / ((sum(x * x for x in a) ** .5) * (sum(y * y for y in b) ** .5) or 1)


def soft_eq(a, b):
    """Lenient equality: normalized-exact OR lexical (difflib) OR semantic (embedding, if cached)."""
    a, b = str(a), str(b)
    if key(a) == key(b):
        return True
    if sim(a.lower(), b.lower()) >= DEF_FUZZ:
        return True
    ea, eb = _emb.get(a), _emb.get(b)
    return ea is not None and eb is not None and cos(ea, eb) >= SEM_FUZZ


def strip_references(text):
    """Drop the reference/bibliography section (its last heading onward)."""
    m = list(re.finditer(r"(?im)^\s*(references|bibliography|works cited)\s*$", text))
    return text[:m[-1].start()] if m else text


# --------------------------------------------------------------------------- fetch + parse
def get_text(oaid, url):
    txt = STEP1 / "parsed" / f"{oaid}.txt"
    if txt.exists():
        return txt.read_text(encoding="utf-8")
    pdf = STEP1 / "pdfs" / f"{oaid}.pdf"
    if not pdf.exists():
        if not url:
            print(f"  [{oaid}] no pdf_url — drop a PDF at {pdf}"); return None
        try:
            ref = re.sub(r"/e?pdf.*$", "", url)          # publisher article page as Referer
            req = urllib.request.Request(url, headers={**UA, "Referer": ref})
            data = urllib.request.urlopen(req, timeout=60).read()
            if not data[:5].startswith(b"%PDF"):
                print(f"  [{oaid}] not a PDF (publisher block) — drop manually at {pdf}"); return None
            pdf.write_bytes(data); print(f"  [{oaid}] fetched {len(data)//1024} KB")
        except Exception as e:                                                    # noqa
            print(f"  [{oaid}] fetch failed ({e}) — drop manually at {pdf}"); return None
    from pypdf import PdfReader
    text = strip_references("\n".join((p.extract_text() or "") for p in PdfReader(str(pdf)).pages))
    txt.write_text(text, encoding="utf-8")
    return text


# --------------------------------------------------------------------------- extract (LLM)
class Metric(BaseModel):
    canonical_name: str                   # reused existing concept name, or a new one if unique
    is_new: bool                          # true only if canonical_name is NOT an existing concept
    raw_description: str                  # the authors' OWN wording/formula, verbatim (~ Comment)
    thematic_type: Optional[str]          # one of THEMATIC_TYPES
    criteria: List[str]                   # qualitative goals it measures (~ EvaluationCriterion)
    indirect_criteria: List[str]          # higher-level goals, e.g. Bikeability
    representation_feature: Optional[str] # one of FEATURES
    measurement_scale: Optional[str]      # one of SCALES
    unit: Optional[str]
    buffer: Optional[str]                 # buffer distance, if measured within one
    aggregate_function: Optional[str]     # e.g. average / sum / min / max
    scoring: Optional[str]                # scoring/normalization function, if any
    weight: Optional[str]                 # weight/importance, if any
    data_source: Optional[str]            # e.g. OpenStreetMap, survey
    parts: List[str]                      # component metrics, if composite


class Extraction(BaseModel):
    index_name: str
    metrics: List[Metric]


PROMPT = """Formalize this bike-network evaluation paper the way a domain expert did for the
VeloNEMO knowledge graph. List every evaluation metric the index uses, as structured records.

What is a metric: the measurable variables an index combines, PLUS named composite measures that are
themselves the unit of analysis (Level of Traffic Stress, relative discomfort, Space Syntax
integration/choice/depth — these are CompositeMetrics, not GraphMetric).
These typically consist of other more atomic metrics.
When such a composite is a CLASSIFICATION defined by thresholds over raw road attributes (road type,
number of lanes, speed limit, slope, residential vs. arterial) — as Level of Traffic Stress is — also
extract each of those underlying attributes as its own metric, since they are the measurable data.

NOT a metric: the paper's
overall index/total score, nor a thematic group or score that simply bundles other metrics (a "Traffic", "Environmental"
or "Infrastructure" component) — those are typically thematic groups to group metrics and its criteria.
When the index is organized into such components, do NOT extract the component name; look INSIDE each
component and extract the specific measurable metrics it is computed from (e.g. bike-lane intersection
density, distance to bike parking, distance to bike-sharing station).

Per metric:
- canonical_name: reuse the EXACT name of an existing concept (list below) when the metric matches
  one BY MEANING, not just wording (e.g. "mixed land use", "land-use entropy" and "Shannon land-use
  index" all -> LandUseMix); else coin a new CamelCase singular name and set is_new=true.
- raw_description: the authors' own wording/formula, verbatim.
- thematic_type: one of {types}. Metrics are typically subclasses of these types (reuse the type the
  ontology already assigns to a metric you are reusing — see the grouped list below).
- criteria: the specific goal(s) it assesses (prefer {criteria}); put umbrella goals (Bikeability,
  Suitability) in indirect_criteria (higher level) instead. 
  In other words, direct criteria are the specific goals the metric measures, while indirect criteria are the broader 
  objectives that the metric contributes to. Both are Criterion concepts. Perhaps indirect is a subclass. 
- representation_feature: the spatial unit the index reports results on — regular grid/hexagon
  cells -> GridCell; an administrative or arbitrary area/zone/buffer -> Area; road segment -> Edge;
  intersection -> Node; whole route -> Route; point location (facility/POI) -> PointOfInterest;
  whole network -> Network. One of {features}.
- measurement_scale: the RAW metric's scale (nominal for categories, ratio for counts/densities/
  distances) — not the index's scoring/rating scale; one of {scales}.
- unit, buffer, aggregate_function, scoring, weight, data_source, parts: fill when the paper states them. When possible,
  reusing existing concepts (e.g. unit = "meter" or "kilometer" from the ontology). 

Infer type/scale when unstated (an expert would). Don't invent or merge metrics.
Existing concepts (reuse exact names where they fit):
{metrics}

PAPER:
{text}
"""


def self_consistency(runs):
    """Keep metrics present in a majority of the K samples; majority-vote each field (D-012)."""
    thr = len(runs) // 2 + 1
    groups = defaultdict(list)
    for run in runs:
        for kk, m in {key(x["canonical_name"]): x for x in run}.items():   # dedup within a sample
            groups[kk].append(m)
    vote = lambda ms, f: Counter(v for m in ms if (v := m.get(f))).most_common(1)
    out = []
    for ms in groups.values():
        if len(ms) < thr:
            continue
        m = dict(ms[0])
        for f in ("canonical_name", "raw_description", "thematic_type",
                  "representation_feature", "measurement_scale", "unit"):
            top = vote(ms, f)
            m[f] = top[0][0] if top else None
        m["is_new"] = sum(bool(x.get("is_new")) for x in ms) > len(ms) / 2
        for f in ("criteria", "indirect_criteria"):
            cc = Counter(c for x in ms for c in (x.get(f) or []))
            m[f] = ([c for c, v in cc.items() if v >= thr] or [cc.most_common(1)[0][0]]) if cc else []
        m["support"] = round(len(ms) / len(runs), 2)
        out.append(m)
    return out


def extract(oaid, text):
    out = STEP1 / "extracted" / f"{oaid}.json"
    if out.exists() and not FORCE_EXTRACT:
        return json.loads(out.read_text(encoding="utf-8"))
    prompt = PROMPT.format(types=THEMATIC_TYPES, criteria=CRITERIA, features=FEATURES,
                           scales=SCALES, metrics=METRIC_VOCAB, text=text)
    gen = types.GenerateContentConfig(response_mime_type="application/json",
                                      response_schema=Extraction, temperature=TEMP)
    runs = [client.models.generate_content(model=MODEL, contents=prompt, config=gen).parsed.model_dump()
            for _ in range(K)]
    data = {"index_name": runs[0].get("index_name", ""),
            "metrics": self_consistency([r["metrics"] for r in runs])}
    out.write_text(json.dumps(data, indent=2), encoding="utf-8")
    print(f"  [{oaid}] {len(data['metrics'])} metrics (majority of {K} samples)")
    return data


# --------------------------------------------------------------------------- ground truth
def load_ground_truth():
    ws = openpyxl.load_workbook(GROUND_TRUTH_XLSX, read_only=True).active
    hdr = {c.value: i for i, c in enumerate(next(ws.iter_rows(max_row=1)))}
    gt = defaultdict(lambda: defaultdict(lambda: {
        "criteria": set(), "parts": set(), "type": None, "scale": None, "feature": None,
        "unit": None, "aggregate": None, "scoring": None, "buffer": None, "comment": ""}))
    for row in ws.iter_rows(min_row=2, values_only=True):
        doi, name = norm_doi(row[hdr["EvaluationMethod"]]), row[hdr["EvaluationMetric"]]
        if not doi or not name:
            continue
        m = gt[doi][name]
        for c in ("EvaluationCriterion", "IndirectEvaluationCriterion"):   # direct ∪ umbrella
            for v in str(row[hdr[c]] or "").split(","):                     # cells may hold "A,B"
                if v.strip():
                    m["criteria"].add(v.strip())
        for v in str(row[hdr["Parts"]] or "").split(","):                  # composite components
            if v.strip():
                m["parts"].add(v.strip())
        m["type"] = m["type"] or row[hdr["MetricType"]]
        m["scale"] = m["scale"] or row[hdr["MeasurementScale"]]
        m["feature"] = m["feature"] or row[hdr["RepresentationFeature"]]
        m["unit"] = m["unit"] or row[hdr["Unit"]]
        m["aggregate"] = m["aggregate"] or row[hdr["Function"]]
        m["scoring"] = m["scoring"] or row[hdr["ScoringFunction"]]
        m["buffer"] = m["buffer"] or row[hdr["Buffer"]]
        m["comment"] = row[hdr["Comment"]] or m["comment"]
    return gt


# --------------------------------------------------------------------------- compare (lenient)
def compare(llm_metrics, gt):
    """Lenient match on name (lexical), definition (lexical), OR definition (semantic embedding)."""
    gt_items, used, pairs = list(gt.items()), set(), []
    embed([lm.get("raw_description") or "" for lm in llm_metrics] +
          [str(v["comment"]) for _, v in gt_items] +          # + field values for soft_eq (agree)
          [str(lm.get(f) or "") for lm in llm_metrics for f in ("unit", "scoring", "aggregate_function")] +
          [str(v.get(f) or "") for _, v in gt_items for f in ("unit", "scoring", "aggregate")])
    for lm in llm_metrics:
        lk, ld = key(lm["canonical_name"]), (lm.get("raw_description") or "")
        le = _emb.get(ld)
        best, best_s = None, 0.0
        for name, info in gt_items:
            if name in used:
                continue
            nl = 1.0 if lk == key(name) else sim(lk, key(name))
            dl = sim(ld.lower(), str(info["comment"]).lower())
            se = cos(le, _emb.get(str(info["comment"])))
            if (nl >= NAME_FUZZ or dl >= DEF_FUZZ or se >= SEM_FUZZ) and max(nl, dl, se) > best_s:
                best, best_s = name, max(nl, dl, se)
        if best:
            used.add(best); pairs.append((lm, best, gt[best], round(best_s, 2)))
    missed = [n for n, _ in gt_items if n not in used]
    extra = [lm for lm in llm_metrics if lm not in [p[0] for p in pairs]]
    return pairs, missed, extra


def agree(pairs):
    """Field agreement over matched pairs, scored ONLY where the human GT annotated the field
    (a blank GT field is 'not annotated', not a disagreement). Returns None if GT never
    annotated that field for this paper."""
    def rate(lm_get, gt_get, overlap=False, eq=lambda a, b: key(a) == key(b)):
        num = den = 0
        for lm, _, g, _ in pairs:
            gv = gt_get(g)
            if not gv:
                continue                                  # GT unannotated → skip
            den += 1
            lv = lm_get(lm)
            if overlap:
                num += bool({key(x) for x in lv} & {key(x) for x in gv})
            else:
                num += bool(lv and eq(lv, gv))
        return num / den if den else None
    num_eq = lambda a, b: re.sub(r"\D", "", str(a)) == re.sub(r"\D", "", str(b))      # buffer distances
    crit = lambda m: (m.get("criteria") or []) + (m.get("indirect_criteria") or [])   # direct ∪ umbrella
    # closed-vocab fields = exact; free-text (unit/aggregate/scoring) = soft_eq; buffer = numeric.
    return {"type":      rate(lambda m: m.get("thematic_type"), lambda g: g["type"]),
            "scale":     rate(lambda m: m.get("measurement_scale"), lambda g: g["scale"]),
            "feature":   rate(lambda m: m.get("representation_feature"), lambda g: g["feature"]),
            "criteria":  rate(crit, lambda g: g["criteria"], overlap=True),
            "unit":      rate(lambda m: m.get("unit"), lambda g: g["unit"], eq=soft_eq),
            "aggregate": rate(lambda m: m.get("aggregate_function"), lambda g: g["aggregate"], eq=soft_eq),
            "scoring":   rate(lambda m: m.get("scoring"), lambda g: g["scoring"], eq=soft_eq),
            "buffer":    rate(lambda m: m.get("buffer"), lambda g: g["buffer"], eq=num_eq),
            "parts":     rate(lambda m: m.get("parts") or [], lambda g: g["parts"], overlap=True)}


# --------------------------------------------------------------------------- report
FIELDS = ("type", "scale", "feature", "criteria")            # per-paper columns
EXT = ("unit", "aggregate", "scoring", "buffer", "parts")    # sparse -> shown as one summary line


def marks(lm, gt):
    """Per-matched-pair field-agreement marks "[T S F C]" = Thematic type · measurement Scale ·
    representation Feature · Criteria.  Symbols: '=' agree · '≠' differ · '·' the human left that
    field blank (so it is not scored)."""
    eq = lambda a, b: "·" if not b else ("=" if a and key(a) == key(b) else "≠")
    lmc = {key(x) for x in (lm.get("criteria") or []) + (lm.get("indirect_criteria") or [])}
    c = "·" if not gt["criteria"] else ("=" if lmc & {key(x) for x in gt["criteria"]} else "≠")
    return (f"T{eq(lm.get('thematic_type'), gt['type'])} S{eq(lm.get('measurement_scale'), gt['scale'])} "
            f"F{eq(lm.get('representation_feature'), gt['feature'])} C{c}")


def diffs(lm, gt):
    """The actual model≠human VALUES behind each '≠' mark, so a low/0% field is traceable to the pair
    that caused it (e.g. a paper with type 0% shows 'type: GraphMetric ≠ CompositeMetric' on each pair)."""
    out = []
    for lab, lv, gv in (("type", lm.get("thematic_type"), gt["type"]),
                        ("scale", lm.get("measurement_scale"), gt["scale"]),
                        ("feature", lm.get("representation_feature"), gt["feature"])):
        if gv and not (lv and key(lv) == key(gv)):
            out.append(f"{lab}: {lv or '—'} ≠ {gv}")
    lmc = {key(x) for x in (lm.get("criteria") or []) + (lm.get("indirect_criteria") or [])}
    if gt["criteria"] and not (lmc & {key(x) for x in gt["criteria"]}):
        out.append(f"criteria: {'/'.join(lm.get('criteria') or []) or '—'} ≠ {'/'.join(sorted(gt['criteria']))}")
    return "  ⟵ " + " · ".join(out) if out else ""


def gt_onto_flag(name, gt):
    """Flag a matched metric whose GT-row thematic type disagrees with the ontology's own subclass."""
    ot = ONTO_TYPE.get(key(name))
    if ot and gt["type"] and key(gt["type"]) != key(ot):
        return f"   ⚠ GT-type `{gt['type']}` ≠ ontology `{ot}`"
    return ""


def write_report(results):
    done = [r for r in results if not r.get("skip")]
    for r in done:
        r["new"] = [lm for lm in [p[0] for p in r["pairs"]] + r["extra"] if lm.get("is_new")]
    pct = lambda v: f"{v:.0%}" if v is not None else "—"

    def field_mean(k):                    # macro-mean of a field's agreement over annotated papers
        vs = [r["agree"][k] for r in done if r["agree"][k] is not None]
        return (sum(vs) / len(vs) if vs else None), len(vs)

    L = ["# Step 1 — LLM vs. human (v1) extraction\n",
         f"_`{MODEL}` · K={K} · match: name/definition (lexical+semantic) · fields scored where GT annotated._\n",
         "| paper | GT | LLM | match | new | recall | prec | supp | " + " | ".join(FIELDS) + " |",
         "|---|--:|--:|--:|--:|--:|--:|--:|" + "--:|" * len(FIELDS)]
    for r in results:
        if r.get("skip"):
            L.append(f"| {r['key']} | | | | | _{r['skip']}_ |" + " |" * (3 + len(FIELDS))); continue
        a = r["agree"]
        L.append(f"| {r['key']} | {r['n_gt']} | {r['n_llm']} | {r['n_match']} | {len(r['new'])} | "
                 f"{r['recall']:.0%} | {r['prec']:.0%} | {(r['supp'] or 0):.2f} | "
                 + " | ".join(pct(a[k]) for k in FIELDS) + " |")
    if done:
        rec, prc = (sum(r[x] for r in done) / len(done) for x in ("recall", "prec"))
        sup = sum(r["supp"] or 0 for r in done) / len(done)
        L.append(f"| **mean** | | | | | **{rec:.0%}** | **{prc:.0%}** | **{sup:.2f}** | "
                 + " | ".join(pct(field_mean(k)[0]) for k in FIELDS) + " |")
        L += ["", "_Extended fields (agreement where the human annotated it): "
              + " · ".join(f"{k} {pct(field_mean(k)[0])} (n={field_mean(k)[1]})" for k in EXT) + "._"]

    L += ["",
          "### Legend",
          "**Table columns**",
          "- **paper** — the source paper (short key).",
          "- **GT** — number of metrics the human annotated for this paper (the ground truth).",
          "- **LLM** — number of metrics the model extracted.",
          "- **match** — model metrics paired to a human one (same metric).",
          "- **new** — metrics the model coined as NOT already in the ontology (`is_new`).",
          "- **recall** — of the human's metrics, the share the model also found (match ÷ GT).",
          "- **prec** — precision: of the model's metrics, the share that matched a human one (match ÷ LLM).",
          "- **supp** — K-run stability: mean `support` over the paper's extracted metrics (1.00 = all K "
          "samples agreed on every metric; 0.67 = on average 2 of 3). Low supp = the extraction itself is "
          "unstable for this paper — read its scores with that uncertainty in mind.",
          "- **type · scale · feature · criteria** — field agreement over matched pairs, model vs. human GT, "
          "scored ONLY where the human annotated that field: thematic **type** (of 6) · measurement **scale** "
          "(of 4) · representation **feature** (of 7) · **criteria** (goals, direct∪indirect — counts as agree "
          "if the sets overlap).",
          "- _Extended fields line (above): the same field-agreement idea for the sparse fields "
          "unit · aggregate · scoring · buffer · parts._",
          "",
          "**Per-pair marks** (in each paper's *Matched* list, e.g. `[T= S≠ F= C=]`)",
          "- `[T S F C]` = agreement on **T**hematic type · measurement **S**cale · representation **F**eature · "
          "**C**riteria.",
          "- `=` agree · `≠` differ · `·` the human left that field blank (so it is not scored).",
          "- `⟵ …` lists the actual differing values (model ≠ human) — this is where a low / 0% score comes from.",
          "- `⚠` flags a pair whose **GT-row type disagrees with the ontology** — a data inconsistency, not a "
          "model error."]

    for r in results:
        if r.get("skip"):
            L += [f"## {r['key']} — skipped ({r['skip']})", ""]; continue
        L += [f"## {r['key']} · _{r['index_name']}_", "",
              "**Matched:**"] + [f"- `{lm['canonical_name']}` → `{g}`  [{marks(lm, gt)}]{diffs(lm, gt)}{gt_onto_flag(g, gt)}"
                                 for lm, g, gt, _ in r["pairs"]]
        if r["missed"]: L.append("**Missed:** " + ", ".join(f"`{x}`" for x in r["missed"]))
        if r["extra"]:  L.append("**Extra:** " + ", ".join(f"`{lm['canonical_name']}`" for lm in r["extra"]))
        if r["new"]:    L.append("**Newly minted:** " + ", ".join(f"`{lm['canonical_name']}`" for lm in r["new"]))
        L.append("")
    (STEP1 / "report" / "step1_report.md").write_text("\n".join(L), encoding="utf-8")

    with (STEP1 / "report" / "matches.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["paper", "status", "llm_metric", "gt_metric", "score", "is_new", "marks",
                    "llm_raw", "gt_comment"])
        for r in done:
            for lm, g, gt, sc in r["pairs"]:
                w.writerow([r["key"], "match", lm["canonical_name"], g, sc, lm.get("is_new"),
                            marks(lm, gt), lm.get("raw_description"), gt["comment"]])
            for x in r["missed"]:
                w.writerow([r["key"], "missed", "", x, "", "", "", "", r["gt"][x]["comment"]])
            for lm in r["extra"]:
                w.writerow([r["key"], "extra", lm["canonical_name"], "", "", lm.get("is_new"), "",
                            lm.get("raw_description"), ""])
    print("\nReport:", STEP1 / "report" / "step1_report.md")


# --------------------------------------------------------------------------- main
def main():
    if not GROUND_TRUTH_XLSX.exists():
        raise SystemExit(f"ground truth not found: {GROUND_TRUTH_XLSX}")
    gt_by_doi = load_ground_truth()
    recs = {}                                     # parse each manifest line once (not 3×)
    for line in MANIFEST_JSONL.open(encoding="utf-8"):
        r = json.loads(line)
        if r["openalex_id"] in SLICE_IDS:
            recs[r["openalex_id"]] = r
    print(f"vocab from ontology: {len(THEMATIC_TYPES)} types, {len(CRITERIA)} criteria, "
          f"{len(FEATURES)} features, {len(SCALES)} scales\n")

    results, last = [], 0.0
    for oaid, k in SLICE_IDS.items():
        print(f"[{k}] {oaid}")
        rec = recs.get(oaid, {})
        gt = gt_by_doi.get(norm_doi(rec.get("doi")), {})
        text = get_text(oaid, rec.get("pdf_url"))
        if not text or len(text) < 500:
            results.append({"key": k, "skip": "no usable text"}); continue
        wait = MIN_INTERVAL - (time.time() - last)
        if wait > 0:
            time.sleep(wait)
        last = time.time()
        try:
            data = extract(oaid, text)
        except Exception as e:                                                   # noqa
            print(f"  [{oaid}] extract error: {e}"); results.append({"key": k, "skip": f"extract error"}); continue
        pairs, missed, extra = compare(data["metrics"], gt)
        ms = data["metrics"]
        results.append({
            "key": k, "index_name": data.get("index_name", ""), "gt": gt,
            "n_gt": len(gt), "n_llm": len(ms), "n_match": len(pairs),
            "recall": len(pairs) / len(gt) if gt else 0.0,
            "prec": len(pairs) / len(ms) if ms else 0.0,
            "supp": sum(m.get("support", 1.0) for m in ms) / len(ms) if ms else None,
            "agree": agree(pairs), "pairs": pairs, "missed": missed, "extra": extra})
        print(f"  GT {len(gt)} · LLM {len(data['metrics'])} · matched {len(pairs)}")
    write_report(results)


if __name__ == "__main__":
    main()
